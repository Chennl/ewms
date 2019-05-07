from flask import render_template,request,current_app,url_for,json,flash,redirect,make_response
from flask_login import current_user,login_required
import random, csv,io,os
from sqlalchemy import and_,text
from datetime import datetime,date
from app.projects import bp
from app import db
from app.models import Project,ProjectMaterial,Material,WarehouseNote,WarehouseNoteItem
from openpyxl import Workbook,load_workbook
import tempfile
       

@bp.route('/index', methods=['GET'])
@login_required
def index():
  role = current_user.can_project()
  return render_template('projects/index.html', title='工程管理',role=role) 


@bp.route('/<int:id>/bom', methods=['GET','POST'])
def bom(id):
  project = Project.query.get_or_404(id)
  materials = ProjectMaterial.query.filter_by(project_id=id)
  return render_template('projects/bom.html', id=id,title='工程需求清单',materials=materials,project=project) 

@bp.route('/<int:id>/edit', methods=['GET'])
def get_project(id):
  project = Project.query.get(id)
  role = current_user.can_project()
  return render_template('projects/projectedit.html',  project=project ,httpaction='PUT',title='修改工程信息',role=role) 
 
@bp.route('/new', methods=['GET'])
def get_new_project():
  project = Project(id=0,name='',year=date.today().year,wbscode='',remark='')
  return render_template('projects/projectedit.html',  project=project ,httpaction='POST',title='添加工程') 


@bp.route('/warehouse', methods=['GET','POST'])
@login_required
def warehouse():
  id = request.args.get('id',1, type=int)
  return render_template('projects/warehouse.html', id=id,title='出入库管理') 

#入库管理
@bp.route('/<int:id>/inbound', methods=['GET'])
def project_inbound(id):
  role = current_user.can_warehouse_in()
  project = Project.query.get(id)
  return render_template('projects/inbound.html', project=project,title='入库管理',role=role) 

@bp.route('/<int:id>/warehousing', methods=['GET'])
def warehousing(id):
  project = Project.query.get(id)
  note = WarehouseNote(id=0,project_id=id,note_no=datetime.now().strftime('%y%m%d%H%M%S'),warehouse_id='',note_date=date.today(),remark='')
  # sbq = db.session.query(WarehouseNoteItem.material_id,WarehouseNoteItem.quantity).filter(WarehouseNoteItem.note_id==note.id).subquery()
  # items = db.session.query(ProjectMaterial,sbq.c.quantity) \
  #         .filter(ProjectMaterial.project_id==note.project_id)\
  #         .outerjoin(sbq, ProjectMaterial.material_id == sbq.c.material_id).all() 
  # data = [{
  #         'id':0,
  #         'material_id':item.ProjectMaterial.material.id,
  #         'material_code':item.ProjectMaterial.material.code,
  #         'material_name':item.ProjectMaterial.material.name,
  #         'material_quantity':item.quantity,
  #         'material_specification':item.ProjectMaterial.material.specification,
  #         'material_category':item.ProjectMaterial.material.category,
  #         'material_unit':item.ProjectMaterial.material.unit,
  #         } for item in items] 
  # return render_template('projects/warehousing.html', note=note,project=project,data=data,title='新增入库单',httpaction="POST") 
  data = get_warehousenote_full_items(project.id,note.id)
  url = '/api/projects/%s/%s/warehousing'%(project.id,note.id)
  return render_template('projects/warehousing.html', note=note,project=project,data=data,title='新增入库单',method="POST",url=url) 

@bp.route('/<int:id>/warehousing/edit', methods=['GET'])
def warehousing_edit(id):
  role = current_user.can_warehouse_in()
  note = WarehouseNote.query.get_or_404(id)
  project = Project.query.get(note.project_id)
  data = get_warehousenote_full_items(project.id,note.id)
  url = '/api/projects/%s/%s/warehousing'%(project.id,note.id)
  return render_template('projects/warehousing.html', note=note,project=project,data=data,title='修改入库单',method="PUT",url=url,role=role) 



#出库管理
@bp.route('/<int:id>/outbound', methods=['GET'])
def project_outbound(id):
  role = current_user.can_warehouse_out()
  project = Project.query.get(id)
  return render_template('projects/outbound.html', project=project,title='出库管理',role=role) 

def get_warehousenote_full_items(project_id,note_id):
  data=[]
  sbq = db.session.query(WarehouseNoteItem.material_id,WarehouseNoteItem.quantity,WarehouseNoteItem.remark).filter(WarehouseNoteItem.note_id==note_id).subquery()
  qy = db.session.query(ProjectMaterial,sbq.c.quantity,sbq.c.remark) \
          .filter(ProjectMaterial.project_id==project_id)\
          .outerjoin(sbq, ProjectMaterial.material_id == sbq.c.material_id)
  items = db.session.query(ProjectMaterial,sbq.c.quantity,sbq.c.remark) \
          .filter(ProjectMaterial.project_id==project_id)\
          .outerjoin(sbq, ProjectMaterial.material_id == sbq.c.material_id).all() 
  if items:
    data = [{
            'id':0,
            'material_id':item.ProjectMaterial.material.id,
            'material_code':item.ProjectMaterial.material.code,
            'material_name':item.ProjectMaterial.material.name,
            'material_quantity':item.quantity,
            'material_specification':item.ProjectMaterial.material.specification,
            'material_category':item.ProjectMaterial.material.category,
            'material_unit':item.ProjectMaterial.material.unit,
            'material_remark':item.remark,
            } for item in items] 
  else:
    data=[]
  
  return data

@bp.route('/<int:id>/warehouseout', methods=['GET'])
def warehouseout(id):
  project = Project.query.get(id)
  note = WarehouseNote(id=0,project_id=id,note_no=datetime.now().strftime('%y%m%d%H%M%S'),warehouse_id='',note_date=date.today(),remark='')
  data = get_warehousenote_full_items(project.id,note.id)
  url = '/api/projects/%s/0/warehouseout'%(project.id)
  return render_template('projects/warehouseout.html', note=note,project=project,data=data,title='新增出库单',method="POST",url=url) 

@bp.route('/<int:id>/warehouseout/edit', methods=['GET'])
def warehouseout_edit(id):
  note = WarehouseNote.query.get_or_404(id)
  project = Project.query.get(note.project_id)
  role = current_user.can_warehouse_out()
  # sbq = db.session.query(WarehouseNoteItem.material_id,WarehouseNoteItem.quantity).filter(WarehouseNoteItem.note_id==note.id).subquery()
  # items = db.session.query(ProjectMaterial,sbq.c.quantity) \
  #         .filter(ProjectMaterial.project_id==note.project_id)\
  #         .outerjoin(sbq, ProjectMaterial.material_id == sbq.c.material_id).all() 
  # if items:
  #   data = [{
  #           'id':0,
  #           'material_id':item.ProjectMaterial.material.id,
  #           'material_code':item.ProjectMaterial.material.code,
  #           'material_name':item.ProjectMaterial.material.name,
  #           'material_quantity':item.quantity,
  #           'material_specification':item.ProjectMaterial.material.specification,
  #           'material_category':item.ProjectMaterial.material.category,
  #           'material_unit':item.ProjectMaterial.material.unit,
  #           'material_remark':item.ProjectMaterial.remark,
  #           } for item in items] 
  # else:
  #   data=[]
  data = get_warehousenote_full_items(project.id,note.id)
  url = '/api/projects/%s/%s/warehouseout'%(project.id,note.id)
  return render_template('projects/warehouseout.html', note=note,project=project,data=data,title='修改出库单',method="PUT",url=url,role=role) 


#退料管理
@bp.route('/<int:id>/return', methods=['GET'])
def project_return(id):
  project = Project.query.get(id)
  role = current_user.can_warehouse_return()
  return render_template('projects/return.html', project=project,title='退料管理',role=role) 
#新增退料单窗口
@bp.route('/<int:id>/warehousenotereturn', methods=['GET'])
def add_warehousenote_return(id):
  project = Project.query.get(id)
  note = WarehouseNote(id=0,project_id=id,note_no=datetime.now().strftime('%y%m%d%H%M%S'),warehouse_id='',note_date=date.today(),remark='')
  # sbq = db.session.query(WarehouseNoteItem.material_id,WarehouseNoteItem.quantity).filter(WarehouseNoteItem.note_id==note.id).subquery()
  # items = db.session.query(ProjectMaterial,sbq.c.quantity) \
  #         .filter(ProjectMaterial.project_id==note.project_id)\
  #         .outerjoin(sbq, ProjectMaterial.material_id == sbq.c.material_id).all() 
  # data = [{
  #         'id':0,
  #         'material_id':item.ProjectMaterial.material.id,
  #         'material_code':item.ProjectMaterial.material.code,
  #         'material_name':item.ProjectMaterial.material.name,
  #         'material_quantity':item.quantity,
  #         'material_specification':item.ProjectMaterial.material.specification,
  #         'material_category':item.ProjectMaterial.material.category,
  #         'material_unit':item.ProjectMaterial.material.unit,
  #         } for item in items] 
  # return render_template('projects/warehousenotereturn.html', note=note,project=project,data=data,title='新增退料单',httpaction="POST") 

  data = get_warehousenote_full_items(project.id,note.id)
  url = '/api/projects/%s/0/warehousereturn'%(project.id)
  return render_template('projects/warehousenotereturn.html', note=note,project=project,data=data,title='新增退料单',method="POST",url=url)

@bp.route('/<int:id>/warehousenotereturn/edit', methods=['GET'])
def update_warehousenote_return(id):
  note = WarehouseNote.query.get_or_404(id)
  role = current_user.can_warehouse_return()
  project = Project.query.get(note.project_id)
  data = get_warehousenote_full_items(project.id,note.id)
  url = '/api/projects/%s/%s/warehousereturn'%(project.id,note.id)
  return render_template('projects/warehousenotereturn.html', note=note,project=project,data=data,title='修改退料单',method="PUT",url=url,role=role)

@bp.route('/<int:id>/inventory')
def get_project_inventory(id):
  project = Project.query.get_or_404(id)
  return render_template('projects/inventory.html', project=project)






# @bp.route('/checkin', methods=['GET','POST'])
# def checkin():
#   id = request.args.get('id',1, type=int)
#   projects = Project.query.all()
#   project = Project.query.get(1)

#   return render_template('projects/bom.html', id=id,title='工程需求清单',materials=project.materials,projects=projects) 

# @bp.route('/search', methods=['GET','POST'])
# def search_materials():
#   id = request.args.get('id',1, type=int)
#   return render_template('projects/searchMaterials.html', id=id,title='工程需求清单') 

@bp.route('/download/csv')
def dowload_project_file():
  projects = Project.query.all()
  sio  = io.StringIO()
  with open("sio",mode='w',newline='',encoding='utf-8') as csvfile:
    fieldnames = ['工程编号', '工程名称', '工程年份','备注','创建日期','修改人员']
    #fieldnames = ['id', 'wbscode', 'name','year','remark','created_date','updated_date','updated_by']
    writer = csv.writer(sio, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)    
    writer.writerow(fieldnames)
    for p in projects:
      writer.writerow([p.id,p.name,p.year,p.remark,p.created_date])
    sio.seek(0) 
    resp = make_response(sio.getvalue())
    resp.headers["Content-Disposition"] = "attachment; filename={}.xlsx".format('name')
    resp.headers['Content-Type'] = 'application/x-xlsx'
    return resp

  return 'download'

@bp.route('/upload_materail')
def dowload_materail_file():
  projects = Project.query.all()
  sio  = io.StringIO()
  with open("sio",mode='w',newline='',encoding='utf-8') as csvfile:
    fieldnames = ['工程编号', '工程名称', '工程年份','备注','创建日期','修改人员']
    #fieldnames = ['id', 'wbscode', 'name','year','remark','created_date','updated_date','updated_by']
    writer = csv.writer(sio, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)    
    writer.writerow(fieldnames)
    for p in projects:
      writer.writerow([p.id,p.name,p.year,p.remark,p.created_date])
    sio.seek(0) 
    resp = make_response(sio.getvalue())
    resp.headers["Content-Disposition"] = "attachment; filename={}.xlsx".format('name')
    resp.headers['Content-Type'] = 'application/x-xlsx'
    return resp

  return 'download'
 

@bp.route('/upload',methods=['GET','POST'])
@login_required
def upload():
  if request.method =='GET':
    return render_template('projects/upload.html')
  elif request.method =='POST':
    return save_upload_file()

def gen_file_name(filename):
  filename = os.path.join(current_app.config['UPLOAD_FOLDER_PATH'],filename)
  i = 1
  while os.path.exists(filename):
      name,extension = os.path.splitext(filename)
      filename = '%s_%s%s'%(name,str(i),extension)
      i += 1
  return filename

@bp.route('/upload_material',methods=['POST'])
def upload_material():
  if request.method =='POST':
    filename = save_upload_file()
    if len(filename)>0:
      import_material_file(filename)
      return '上传成功'
    else:
      return '上传失败'

@bp.route('/upload_bom',methods=['POST'])
def upload_bom():
  if request.method =='POST':
    filename = save_upload_file()
    if len(filename)>0:
      logger_msg = import_bom_file(filename)
      return ''.join('<p>%s'%(v) for v in logger_msg)
    else:
      return '上传失败'

@bp.route('/upload_inbound',methods=['POST'])
def upload_inbound():
  if request.method =='POST':
    filename = save_upload_file()
    if len(filename)>0:
      logger_msg = import_warehouseItem_file(filename)
      return ''.join('<p>%s'%(v) for v in logger_msg)
    else:
      return '上传未完成失败'.join('<p>%s'%(v) for v in logger_msg)




@bp.route('/upload_project',methods=['POST'])
def upload_project():
  if request.method =='POST':
    filename = save_upload_file()
    if len(filename)>0:
      logger_msg = import_bom_file(filename)
      return logger_msg
    else:
      return '上传失败'

def save_upload_file():
    print(request.files)
    files = request.files['file']
    if files:
       filename = files.filename
       filename = gen_file_name(filename)
       mime_type = files.content_type
       upload_file_path = os.path.join(current_app.config['UPLOAD_FOLDER_PATH'],filename)
       try:
         files.save(filename)
       except IOError:
         current_app.logger.exception("fail to save file: %s"%upload_file_path)
         return ''
       return filename
    else:
        return ''

def import_material_file(filename):
  wb = load_workbook(filename)    #打开excel表格
  sheets = wb.get_sheet_names()  # 获取所有的表格
  sheets_first = sheets[0]    # 获取第一个表
  sheet = wb.get_sheet_by_name(sheets_first)  #根据工作表名获取工作表
  i = 0 
  objects=[]
  fileds=['id','code','name','category','specification','unit','','']
  row_cout=sheet.rows
  for row in sheet.rows:
    i += 1
    id=row[0].value
    category = row[1].value
    code =row[2].value
    name =row[3].value
    specification = row[4].value
    unit = row[5].value
    m= Material(id=id,category=category,code=code,name=name,specification=specification,unit=unit,updated_by='system')
    objects.append(m)
    if (i+1)%5 == 0:
      db.session.add_all(objects)
      db.session.commit()
      objects.clear()
    elif (i+1) == sheet.max_row:
      db.session.add_all(objects)
      db.session.commit()
      objects.clear()

 
def import_warehousenote_file_TBC(filename):
  logmsg=[]
  wb = load_workbook(filename)    #打开excel表格
  sheets = wb.get_sheet_names()  # 获取所有的表格
  sheets_first = sheets[0]    # 获取第一个表
  sheet = wb.get_sheet_by_name(sheets_first)  #根据工作表名获取工作表
  
  for sheet in wb.worksheets:

      if sheet.title.startswith('材料ID')  or sheet.title.startswith('工程ID'):
         logmsg.append('跳过工作页[{}]***************.'.format(sheet.title))
         continue

      i = 0 
      objects=[]
      fileds=['project_id','material_id','quantity']
      note_id= None
      sheet_index=1

      #logmsg.append('******************开始导入[{}]共{}行记录***************.'.format(sheet.title,sheet.max_row))

      for row in sheet.rows:
        i += 1
        if i==1:
          continue

        project_id=row[0].value
        material_id = row[1].value
        quantity =row[2].value

        if project_id is None or material_id is None:
          logmsg.append('[{}] 第 {} 行数据不正确.'.format(sheet.title,i))
          continue

        if note_id is None:
          warehouseNote = WarehouseNote.query.filter_by(project_id=project_id).first()
          if warehouseNote is None:
             logmsg.append('[{}] 第 {} 行数据不正确.工程{}找不到对应的入库单'.format(sheet.title,i,project_id))
             break
          note_id = warehouseNote.id
        m= WarehouseNoteItem(note_id=note_id,material_id=material_id,quantity=quantity)
        objects.append(m)

      try:
        
        warehouseNote = WarehouseNote()
        #(	101,'44','190419001'	,'default','入库单',curdate(),'',now(),'sys');
        sql = 'DELETE FROM  warehousenoteitem WHERE note_id=:note_id'
        db.engine.execute(text(sql), {'note_id':note_id})
        if len(objects)>0:
          db.session.add_all(objects)
          db.session.commit()
      except Exception as error:
          current_app.logger.debug('exception is %s' % error)
      finally:
        objects.clear()
        note_id= None

      logmsg.append('工程[{}]共导入{}/{}行记录.'.format(sheet.title,i,sheet.max_row))

  return logmsg

def import_warehouseItem_file(filename):
  logmsg=[]
  wb = load_workbook(filename)    #打开excel表格
  sheets = wb.get_sheet_names()  # 获取所有的表格
  sheets_first = sheets[0]    # 获取第一个表
  sheet = wb.get_sheet_by_name(sheets_first)  #根据工作表名获取工作表
  
  for sheet in wb.worksheets:

      if sheet.title.startswith('材料ID')  or sheet.title.startswith('工程ID'):
         logmsg.append('跳过工作页[{}]***************.'.format(sheet.title))
         continue

      i = 0 
      objects=[]
      fileds=['project_id','material_id','quantity']
      note_id= None
      sheet_index=1

      #logmsg.append('******************开始导入[{}]共{}行记录***************.'.format(sheet.title,sheet.max_row))

      for row in sheet.rows:
        i += 1
        if i==1:
          continue

        project_id=row[0].value
        material_id = row[1].value
        quantity =row[2].value

        if project_id is None or material_id is None:
          logmsg.append('[{}] 第 {} 行数据不正确.'.format(sheet.title,i))
          continue

        if note_id is None:
          warehouseNote = WarehouseNote.query.filter_by(project_id=project_id).first()
          if warehouseNote is None:
             logmsg.append('[{}] 第 {} 行数据不正确.工程{}找不到对应的入库单'.format(sheet.title,i,project_id))
             break
          note_id = warehouseNote.id
        m= WarehouseNoteItem(note_id=note_id,material_id=material_id,quantity=quantity)
        objects.append(m)

      try:
        sql = 'DELETE FROM  warehousenoteitem WHERE note_id=:note_id'
        db.engine.execute(text(sql), {'note_id':note_id})
        if len(objects)>0:
          db.session.add_all(objects)
          db.session.commit()
      except Exception as error:
          current_app.logger.debug('exception is %s' % error)
      finally:
        objects.clear()
        note_id= None

      logmsg.append('工程[{}]共导入{}/{}行记录.'.format(sheet.title,i,sheet.max_row))

  return logmsg

 
def import_bom_file(filename):
  logmsg=[]
  wb = load_workbook(filename)    #打开excel表格
  sheets = wb.get_sheet_names()  # 获取所有的表格
  sheets_first = sheets[0]    # 获取第一个表
  sheet = wb.get_sheet_by_name(sheets_first)  #根据工作表名获取工作表
  
  for sheet in wb.worksheets:
      i = 0 
      objects=[]
      fileds=['id','project_id','material_id','quantity']
     
      #logmsg.append('******************开始导入[{}]共{}行记录***************.'.format(sheet.title,sheet.max_row))
      for row in sheet.rows:
        i += 1
        if i==1:
          continue
        project_id=row[0].value
        material_id = row[1].value
        quantity =row[2].value
        if project_id is None or material_id is None:
          logmsg.append('[{}] 第 {} 行数据不正确.'.format(sheet.title,i))
          continue
        m= ProjectMaterial(project_id=project_id,material_id=material_id,quantity=quantity,updated_by='system')
        objects.append(m)
      try:
        sql = 'DELETE FROM  project_material WHERE project_id=:project_id'
        db.engine.execute(text(sql), {'project_id':project_id})
        if len(objects)>0:
          db.session.add_all(objects)
          db.session.commit()
      except Exception as error:
          current_app.logger.debug('exception is %s' % error)
      finally:
        objects.clear()

      logmsg.append(' 工程材料需求清单[{}]，共导入{}/{}行记录 .'.format(sheet.title,i,sheet.max_row))

  return logmsg

def import_project_file(filename):
  wb = load_workbook(filename)    #打开excel表格
  sheets = wb.get_sheet_names()  # 获取所有的表格
  sheets_first = sheets[0]    # 获取第一个表
  sheet = wb.get_sheet_by_name(sheets_first)  #根据工作表名获取工作表
  i = 0 
  objects=[]
  fileds=['id','name','wbscode','year','area','category','reserve_code','remark']
 
  row_cout=sheet.rows
  for row in sheet.rows:
    i += 1
    id=row[0].value
    name = row[1].value
    wbscode =row[2].value
    year =row[3].value
    area = row[4].value
    category = row[5].value
    reserve_code = row[6].value
    remark = row[8].value

    # project = Project.get(id)
    # m= Material(id=id,category=category,code=code,name=name,specification=specification,unit=unit,updated_by='system')

    # objects.append(m)
    # if (i+1)%5 == 0:
    #   db.session.add_all(objects)
    #   db.session.commit()
    #   objects.clear()
    # elif (i+1) == sheet.max_row:
    #   db.session.add_all(objects)
    #   db.session.commit()
    #   objects.clear()

def import_material_file(filename):
  wb = load_workbook(filename)    #打开excel表格
  sheets = wb.get_sheet_names()  # 获取所有的表格
  sheets_first = sheets[0]    # 获取第一个表
  sheet = wb.get_sheet_by_name(sheets_first)  #根据工作表名获取工作表
  i = 0 
  objects=[]
  fileds=['id','code','name','category','specification','unit','','']
  row_cout=sheet.rows
  for row in sheet.rows:
    i += 1
    id=row[0].value
    category = row[1].value
    code =row[2].value
    name =row[3].value
    specification = row[4].value
    unit = row[5].value
    m= Material(id=id,category=category,code=code,name=name,specification=specification,unit=unit,updated_by='system')
    objects.append(m)
    if (i+1)%5 == 0:
      db.session.add_all(objects)
      db.session.commit()
      objects.clear()
    elif (i+1) == sheet.max_row:
      db.session.add_all(objects)
      db.session.commit()
      objects.clear()



@bp.route('/download/material',methods=['GET','POST'])
@login_required
def export_material_file():
  wb =  Workbook()
  ws  = wb.active
  try:
    ws.title='材料清单'
    fieldnames = ['ID','编号', '分类','名称', '规格型号','单位','修改人员','更新日期']
    row = 1
    col = 1
    for name in fieldnames:
      ws.cell(column=col,row=row,value=name)
      col += 1
    result = Material.query.all()
    for m in result:
      row += 1
      ws.cell(column=1,row=row,value=m.id)
      ws.cell(column=2,row=row,value=m.code)
      ws.cell(column=3,row=row,value=m.category)
      ws.cell(column=4,row=row,value=m.name)
      ws.cell(column=5,row=row,value=m.specification)
      ws.cell(column=6,row=row,value=m.unit)
      ws.cell(column=7,row=row,value=m.updated_by)
      ws.cell(column=8,row=row,value=m.updated_date)
      
    out = io.BytesIO()
    wb.save(out)
    out.seek(0) 
    resp = make_response(out.getvalue())

    resp.headers["Content-Disposition"] = "attachment; filename={}.xlsx".format('material-list-%s'%(date.today().isoformat()))
    resp.headers['Content-Type'] = 'application/x-xlsx'
    return resp
  except Exception:
     current_app.logger.exception("fail to export file:")
     return '导出失败'


@bp.route('/download/project',methods=['GET','POST'])
@login_required
def export_project_file():
  wb =  Workbook()
  ws  = wb.active
  try:
    ws.title='工程列表'
    column_names = ['ID','工程编号', '工程名称', '工程年份','备注','修改人员','更新日期']
    ws.append(column_names)
    row = 2
    col = 1

    result = Project.query.all()
    for m in result:
      ws.cell(column=1,row=row,value=m.id)
      ws.cell(column=2,row=row,value=m.wbscode)
      ws.cell(column=3,row=row,value=m.name)
      ws.cell(column=4,row=row,value=m.year)
      ws.cell(column=5,row=row,value=m.remark)
      ws.cell(column=6,row=row,value=m.updated_by)
      ws.cell(column=7,row=row,value=m.updated_date)
      row += 1

    ws2 = wb.create_sheet(title="工程材料清单")
    column_names = ['ID','工程ID','材料ID','材料编号','材料名称','数量','单位']
    ws2.append(column_names)
    result = ProjectMaterial.query.all()
    row=2
    for m in result:
      ws2.cell(column=1,row=row,value=m.id)
      ws2.cell(column=2,row=row,value=m.project_id)
      ws2.cell(column=3,row=row,value=m.material_id)
      ws2.cell(column=4,row=row,value=m.material.code)
      ws2.cell(column=5,row=row,value=m.material.name)
      ws2.cell(column=6,row=row,value=m.quantity)
      ws2.cell(column=7,row=row,value=m.material.unit)
      row += 1
   

    out = io.BytesIO()
    wb.save(out)
    out.seek(0) 
    resp = make_response(out.getvalue())

    resp.headers["Content-Disposition"] = "attachment; filename={}.xlsx".format('project-list-%s'%(date.today().isoformat()))
    resp.headers['Content-Type'] = 'application/x-xlsx'
    return resp
  except Exception:
     current_app.logger.exception("fail to export file:")
     return '工程导出失败'




@bp.route('/download/warehousenote',methods=['GET','POST'])
@login_required
def export_warehousenote_file():
  wb =  Workbook()
  ws  = wb.active
  try:
    ws.title='仓库单据'
    column_names = ['工程ID','单据ID', '单据编号','单据类型','仓库', '日期','备注','操作人员','操作日期']
    ws.append(column_names)
    row = 2
    result = WarehouseNote.query.all()
 
    for m in result:
      ws.cell(column=1,row=row,value=m.project_id)
      ws.cell(column=2,row=row,value=m.id)
      ws.cell(column=3,row=row,value=m.note_no)
      ws.cell(column=4,row=row,value=m.note_type)
      ws.cell(column=5,row=row,value=m.warehouse_id)
      ws.cell(column=6,row=row,value=m.note_date)
      ws.cell(column=7,row=row,value=m.remark)
      ws.cell(column=8,row=row,value=m.updated_by)
      ws.cell(column=9,row=row,value=m.updated_date)
      row += 1


        


    ws2 = wb.create_sheet(title="单据明细")
    column_names = ['ID','单据ID','材料ID','材料编号','材料名称','数量','单位']
    ws2.append(column_names)
    result = WarehouseNoteItem.query.all()
    row=2
    for m in result:
      ws2.cell(column=1,row=row,value=m.id)
      ws2.cell(column=2,row=row,value=m.note_id)
      ws2.cell(column=3,row=row,value=m.material_id)
      ws2.cell(column=4,row=row,value=m.material.code)
      ws2.cell(column=5,row=row,value=m.material.name)
      ws2.cell(column=6,row=row,value=m.quantity)
      ws2.cell(column=7,row=row,value=m.material.unit)
      row += 1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

     

    resp = make_response(out.getvalue())

    resp.headers["Content-Disposition"] = "attachment; filename={}.xlsx".format('warehousenote-%s'%(date.today().isoformat()))
    resp.headers['Content-Type'] = 'application/x-xlsx'
    return resp
  except Exception:
     current_app.logger.exception("fail to export file:")
     return '仓库单据导出失败'


@bp.route('/download',methods=['GET'])
@login_required
def download():
 return render_template('projects/download.html', title='数据下载') 




 

 

